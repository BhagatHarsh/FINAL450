import openpyxl
import os
from string import punctuation

template = """
# Link: %s
for loop in range(int(input())):
    n,m = map(int,input().split())
    l = list(map(int,input().split()))
"""


def convertCase(
    arr
):  # https://www.geeksforgeeks.org/convert-given-array-of-strings-to-a-camel-case-format-sentence/
    ans = ""
    try:
        N = len(arr)
        for i in range(N):
            if (len(ans) > 0):
                ans += ' '
            ans += arr[i][0].upper()
            j = 1
            while j < len(arr[i]):
                if (arr[i][j] == ' '):
                    ans += ' '
                    ans += arr[i][j + 1].upper()
                    j += 1
                else:
                    ans += arr[i][j].lower()
                j += 1
    except:
        pass
    return ans.translate(str.maketrans('', '', punctuation))


def getNthLink(n: int):
    nextFile = n-1
    i = 6
    val = 0
    while (val != nextFile):
        i += 1
        if (ws.cell(row=i, column=2).value != None):
            val += 1
        
    requiredCell = ws.cell(row=i, column=2)
    
    nextName = str(nextFile + 1) + '_' + convertCase(requiredCell.value.split()) + ".py"
    cellLink = requiredCell.hyperlink.target

    if (input("Do you want to overwrite the file? (y/n)") == 'y'):
        with open(os.getcwd() + '/FINAL450/' + nextName.replace(" ", ""), 'w') as f:
            f.write(template % (str(cellLink)))
    print("catergory: " + ws.cell(row=i, column=1).value)
    return cellLink


def getNextLink():
    nextFile = 0
    nextName = ''
    for file in os.listdir(os.getcwd() + '/FINAL450/'):
        try:
            nextFile = max(nextFile, int(file.split('_')[0]))
        except:
            pass
    print("getting number " + str(nextFile + 1) + " for you!!")

    i = 6
    val = 0
    while (val != nextFile):
        i += 1
        if (ws.cell(row=i, column=2).value != None):
            val += 1
        
    requiredCell = ws.cell(row=i, column=2)
    
    nextName = str(nextFile + 1) + '_' + convertCase(requiredCell.value.split()) + ".py"
    cellLink = requiredCell.hyperlink.target

    if (input("Do you want to overwrite the file? (y/n)") == 'y'):
        with open(os.getcwd() + '/FINAL450/' + nextName.replace(" ", ""), 'w') as f:
            f.write(template % (str(cellLink)))
    print("catergory: " + ws.cell(row=i, column=1).value)
    return cellLink


wb = openpyxl.load_workbook(os.getcwd() + '/FINAL450/0_FINAL450.xlsx')
ws = wb['Sheet1']

if (input("Do you want the next link? (y/n)") == 'y'):
    print(getNextLink())

elif (input("Do you want a specific link? (y/n)") == 'y'):
    print(getNthLink(int(input("Enter the number of the question: "))))

# This will fail if there is no hyperlink to target
